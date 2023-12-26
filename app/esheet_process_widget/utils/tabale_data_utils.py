from datetime import datetime
from loguru import logger
from typing import List

from app.esheet_process_widget.epw_define import ExcelDataTableWidgetItemDataStruct, SameYTCountTableWidgetItemDataStruct
from openpyxl import load_workbook
from collections import Counter

from app.utils.dev_config import YtBindDevConfigGenerate
from app.utils.tools import is_instance_variables_has_empty


class FileContentIsEmptyException(Exception):
    """
    自定义异常类，用于标记要处理的文件是空的情况
    """

    def __init__(self, message="要处理的数据文件为空"):
        self.message = message
        super().__init__(self.message)


def getExcelDataTableWidgetData(filePath: str, shipid_CID: str = None, scantime_CID: str = None, ytname_CID: str = None, scantimeformat: str = None) -> List[ExcelDataTableWidgetItemDataStruct]:
    # 返回excel表格中指定列的数据

    # [[默认格式]]
    # 单号列 = A
    # 月台号列 = AB
    # 扫描时间列 = AD
    # 扫描时间格式 = %Y-%m-%d %H:%M:%S
    shipID_CID = "A" if shipid_CID is None else shipid_CID
    scanTime_CID = "AD" if scantime_CID is None else scantime_CID
    ytName_CID = "AB" if ytname_CID is None else ytname_CID
    scanTimeFormat = "%Y-%m-%d %H:%M:%S" if scantimeformat is None else scantimeformat

    wb = load_workbook(filePath)

    ws = wb.active  # 获取当前活跃的sheet,默认是第一个sheet
    logInfoText = "\n" + shipID_CID + "列表头为:" + str(ws[shipID_CID][0].value) \
                  + "\n" + scanTime_CID + "列表头为:" + str(ws[scanTime_CID][0].value) \
                  + "\n" + ytName_CID + "列表头为:" + str(ws[ytName_CID][0].value)
    logger.info(logInfoText + "\n注意检查表头数据是否正确")

    ws.delete_rows(1)  # 删除第一行标题
    shipIDList = [cell.value for cell in ws[shipID_CID]]  # 获取单号列数据
    scanTimeList = [cell.value for cell in ws[scanTime_CID]]  # 扫描时间列数据
    ytNameList = [cell.value for cell in ws[ytName_CID]]  # 月台号列数据
    wb.close()  # 用完立即就关掉
    edtw_ItemDataList = []

    for index in range(len(shipIDList)):
        errorTimes = 0  # 报错计数
        temp_ItemData = ExcelDataTableWidgetItemDataStruct()
        try:
            temp_ItemData.shipID = int(shipIDList[index]) if shipIDList[index] is not None else None
            temp_ItemData.scanTime = datetime.strptime(scanTimeList[index], scanTimeFormat) if scanTimeList[index] is not None else None
            temp_ItemData.ytName = str(ytNameList[index])  # 某列数据里可能是空的
            if is_instance_variables_has_empty(temp_ItemData):
                continue
            else:
                edtw_ItemDataList.append(temp_ItemData)
        except IndexError as errorInfo:  # 某个单元格为空，但应该不会存在这个问题的，先抓了再说吧
            errorTimes += 1
            logger.error(f"可能是某个单元格为空，原报错信息{errorInfo}")
            continue
        except ValueError as errorInfo:  # 一般是用户填错了列号
            errorTimes += 1
            logger.error(f"应该是填错了数据列号或者时间格式不对，原报错信息{errorInfo}")
            continue
        if errorTimes >= 100:  # 都错100次了，肯定传错啥了，别处理了
            logger.error(f"错误次数达到100次,函数退出执行!")
            break
    if not edtw_ItemDataList:  # 判断列表是否为空
        raise FileContentIsEmptyException(f"要处理的文件是空的，注意检查是否填错列号\n{filePath}")
    else:
        return edtw_ItemDataList


def getSameYTCountTableWidgetData(devConfigGenerate: YtBindDevConfigGenerate, edtw_ItemDataList: List[ExcelDataTableWidgetItemDataStruct]) -> List[SameYTCountTableWidgetItemDataStruct]:
    # 处理getExcelDataTableWidgetData返回的数据
    # 得到 相同月台的单号总量和月台通道配置

    ytNameList = [edtw_ItemData.ytName for edtw_ItemData in edtw_ItemDataList]
    counter = Counter(ytNameList)  # 字典子类用于统计可哈希对象数量
    sameYTCount = counter.most_common()  # 最大值顺序

    # 这里拿到了单号总量和月台通道配置，应该新开一个list

    sytctw_ItemDataList = []
    for yt, shipCount in sameYTCount:
        temp_ItemDataList = SameYTCountTableWidgetItemDataStruct()
        temp_ItemDataList.ytName = yt
        temp_ItemDataList.shipCount = shipCount
        devconfig = devConfigGenerate.send(yt)  # 向生成器发送月台名称,生成器会返回设备配置
        temp_ItemDataList.devChannel = f"{devconfig.devIP}-{str(devconfig.devChannel)}" if devconfig else "未配置"
        next(devConfigGenerate)  # 驱动生成器执行到ytName = yield，以便接收下一个月台名称
        sytctw_ItemDataList.append(temp_ItemDataList)
    return sytctw_ItemDataList


if __name__ == "__main__":
    # print(getYtConfigFromConfigFile(__devConfigFilePath))
    pass
