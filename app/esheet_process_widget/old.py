import sys
import os
from PyQt5.QtCore import Qt, pyqtSlot
from PyQt5.QtWidgets import (
    QApplication,
    QFileDialog,
    QMessageBox,
    QTableWidgetItem,
    QWidget, QListWidgetItem,
)


# TODO 我的装饰器没写出来
# TODO 拖拽功能也没写
# TODO closeevent保存页面参数

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


class QmyExcelProcess(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)  # 调用父类构造函数，创建窗体
        self.ui = Ui_EPWC()  # 创建UI对象
        self.ui.setupUi(self)  # 构造UI界面
        self.setAcceptDrops(True)  # 允许放置操作
        #

    # ------------------ 以下是VDPw主要调用的方法

    def getAllOfExcelFileData(self):
        # 第一列是文件名，然后依次就是ship，扫描时间，月台 返回所有文件的数据列表
        # 应该是要出一个bug  就是只有在项目切换的时候才会重新冲刷组件的data
        # 然后这个方法直接获取的数据就会有问题
        allOfExcelFileData = []
        for index in range(self.ui.excelFileLW.count()):
            item = self.ui.excelFileLW.item(index)  # 便利所有行的item的data
            itemData = item.data(Qt.UserRole)  # 获取项目
            tableData = itemData['sortedData']  # 从项目中获取表格数据
            fileAddress = itemData['fileAddress']  #
            fileBaseName = os.path.basename(fileAddress)  # 获取文件名
            fileNameSplit = os.path.splitext(fileBaseName)  # 切割文件名和文件扩展名
            fileName = fileNameSplit[0]  # 仅获取文件名
            for rowList in tableData:  # 便利表格数据的每一个行
                rowList.insert(0, str(fileName))  # 每行的列表的第一个索引插入文件名
                allOfExcelFileData.append(rowList)  # 加入大一统
        return allOfExcelFileData

    # ------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------
    # ------------------------------------------------------------------------------------

    @pyqtSlot()
    def on_reprocessExcelFilePB_clicked(self):
        # 重新处理文件按钮被点击
        fileAddress = self.getExcelFileAddressLEText()
        prettyListData = self.handleExcelFileToPretty(fileAddress)  # 取文件名接着进行处理
        self.replaceExcelFileLWItemData(prettyListData)

    def replaceExcelFileLWItemData(self, DictData):
        # 替换ExcelFile列表组件上的项目的数据 也就是保存的作用
        selectedItems = self.ui.excelFileLW.selectedItems()
        selectedItemRow = self.ui.excelFileLW.row(selectedItems[0])  # 获取List组件中选中的第一个项目(虽然是单选模式)
        print("将要删除ExcelFileLW中第" + str(selectedItemRow) + "行的Data")
        selectedItem = self.ui.excelFileLW.takeItem(selectedItemRow)
        selectedItem.setData(Qt.UserRole, DictData)  # 将pop出的listitem设置上新处理的数据
        self.ui.excelFileLW.insertItem(selectedItemRow, selectedItem)  # 从删除的行位置重新插入一个
        print("将要在ExcelFIleLW中的第" + str(selectedItemRow) + "行插入" + str(selectedItem) + "数据")
        self.ui.excelFileLW.setCurrentItem(selectedItem)  # 然后设置选中项
        print("设置ExcelFileLW的当前项目为:" + str(selectedItem))

    @pyqtSlot()
    def on_deleteExcelFileLWItemPB_clicked(self):
        # 删除选中文件按钮点击
        selectedItems = self.ui.excelFileLW.selectedItems()  # 获取List组件中选中的项目
        if selectedItems:
            selectedItemRow = self.ui.excelFileLW.row(selectedItems[0])
            print("将要删除ExcelFileLW中第" + str(selectedItemRow) + "行的Data")
            self.ui.excelFileLW.takeItem(selectedItemRow)

    @pyqtSlot()
    def on_keepShipPB_clicked(self):
        excelTableData = self.getDataInExcelDataTW()  # 获取ExceDataTableWidget上的二维数据
        keepShipNum = int(self.ui.keepShipNumCB.currentText())
        if len(excelTableData) > keepShipNum:
            ytData = [ytStr[2] for ytStr in excelTableData]
            ytSet = set()  # 月台集合
            [ytSet.add(str(i)) for i in ytData]  # 收集月台集合
            excelTableData.sort(key=lambda elem: elem[2])
            sortedYTData = excelTableData[:]
            ytSet = sorted(list(ytSet))  # 转换月台集合为列表并排序
            sameYTData = [[] for i in range(len(ytSet))]
            [
                sameYTData[j].append(sortedYTData[i])
                for j in range(len(ytSet))
                for i in range(len(sortedYTData))
                if sortedYTData[i][2] == ytSet[j]
            ]
            sameYTData.sort(key=lambda x: len(x), reverse=True)  # 列表按照元素长度进行倒序排列
            for sameYTIndex, sameYTList in enumerate(sameYTData):
                ListLength = len(sameYTList)  # 取当前列表的长度
                if ListLength > keepShipNum:  # 如果超出大小 直接一个超出数量的pop循环就可以了
                    for i in range(ListLength - keepShipNum):
                        sameYTData[sameYTIndex].pop()
                else:  # 因为我是进行过排序的 如果有一个不大于目标值那剩下的就更不可能了
                    break
            excelTableData = [j for i in sameYTData for j in i]
            self.setInExcelDataTWPrettyDataToSameYTTW_ExcelDataTW_ExcelFIleAddressLE_exceptYTCBs_onlyYTCBs(
                excelTableData)

    @pyqtSlot()
    def on_deleteExcelDataTWRowPB_clicked(self):
        # 删除行按钮点击 先把table取下来 然后对列表的指定行进行删除 总共得到三个数据再放到各组件上
        excelTableData = self.getDataInExcelDataTW()  # 获取ExceDataTableWidget上的二维数据
        selectedItems = self.ui.excelDataTW.selectedItems()  # 获取表格组件中选中的项目
        if len(selectedItems) == 0:  # 如果没选
            return
        selectedRows = [selectedItem.row() for selectedItem in selectedItems]  # 获取所有选中项目的行
        sortedSelectedRows = sorted(list(set(selectedRows)), reverse=True)  # 先去重，再转列表，最后倒序排列
        print("将要删除ExcelDataTW上的" + str(sortedSelectedRows) + "的数据行")
        for i in range(len(excelTableData) - 1, -1, -1):  # 倒序循环
            if i in sortedSelectedRows:  # 如果当前索引在欲删除索引列表中
                excelTableData.pop(i)  # pop出指定索引数据行
        self.setInExcelDataTWPrettyDataToSameYTTW_ExcelDataTW_ExcelFIleAddressLE_exceptYTCBs_onlyYTCBs(excelTableData)

    @pyqtSlot()
    def on_onlyDeletePB_clicked(self):
        # 只删除按钮点击 先把table取下来 然后重新对列表的月台列进行删除 总共得到三个数据再放到各组件上
        # 再有机会写表格 一定要定义数据列的枚举索引值
        onlyYTCBText = []
        for CBWidget in [self.ui.onlyYT1CB, self.ui.onlyYT2CB, self.ui.onlyYT3CB]:
            onlyYTCBText.append(CBWidget.currentText())  # 循环获取三个组件上的文本
        excelTableData = self.getDataInExcelDataTW()  # 获取ExceDataTableWidget上的二维数据
        for i in range(len(excelTableData) - 1, -1, -1):  # 倒叙循环删除
            rowText = excelTableData[i][2]
            if rowText in onlyYTCBText:  # 如果当前列表的月台文本不在那个三个组件组成的文本列表中
                excelTableData.pop(i)  # 就用pop删除当前索引的列表数据
        print("将要删除ExcelDataTW上月台为" + str(onlyYTCBText) + "的数据行")
        self.setInExcelDataTWPrettyDataToSameYTTW_ExcelDataTW_ExcelFIleAddressLE_exceptYTCBs_onlyYTCBs(excelTableData)

    @pyqtSlot()
    def on_exceptDeletePB_clicked(self):
        # 除了。。。按钮点击 和onlyDelete只删除的处理逻辑只有判断运算符相反
        exceptYTCBText = []
        for CBWidget in [self.ui.exceptYT1CB, self.ui.exceptYT2CB, self.ui.exceptYT3CB]:
            exceptYTCBText.append(CBWidget.currentText())  # 循环获取三个组件上的文本
        excelTableData = self.getDataInExcelDataTW()  # 获取ExceDataTableWidget上的二维数据
        for i in range(len(excelTableData) - 1, -1, -1):  # 倒叙循环删除
            rowText = excelTableData[i][2]
            if rowText not in exceptYTCBText:  # 如果当前列表的月台文本不在那个三个组件组成的文本列表中
                excelTableData.pop(i)  # 就用pop删除当前索引的列表数据
        print("ExcelDataTW上的月台数据行除了" + str(exceptYTCBText) + ",其它都要删除")
        self.setInExcelDataTWPrettyDataToSameYTTW_ExcelDataTW_ExcelFIleAddressLE_exceptYTCBs_onlyYTCBs(excelTableData)

    def setInExcelDataTWPrettyDataToSameYTTW_ExcelDataTW_ExcelFIleAddressLE_exceptYTCBs_onlyYTCBs(self, excelTableData):
        # 设置从ExcelDataTW上取下来的漂亮数据到三大组件和六小组件 这个函数存在的原因是这不是第一次处理文件数据所以无需对文件地址进行操作
        # 其次是这个取放的操作非常一致，所以就暂时给个函数拼接一下也没什么问题
        sortedData, sameYTCount, ytSet = self.handleExcelTwoDimensionalData(excelTableData)
        fileAddress = self.getExcelFileAddressLEText()  # TODO 别奇怪 为什么只是对数据的处理 却还要访问文件地址 因为函数参数的规定
        sortedData_sameYTCount_fileAddress_ytSet = {"sortedData": sortedData, "sameYTCount": sameYTCount,
                                                    "ytSet": ytSet, "fileAddress": fileAddress}
        self.replaceExcelFileLWItemData(sortedData_sameYTCount_fileAddress_ytSet)
        self.setPrettyDataToSameYTTW_ExcelDataTW_ExcelFIleAddressLE_exceptYTCBs_onlyYTCBs(
            sortedData_sameYTCount_fileAddress_ytSet)

    @pyqtSlot(QListWidgetItem, QListWidgetItem)
    def on_excelFileLW_currentItemChanged(self, current, previous):
        # 项目切换 先保存当前界面四个组件的数据 然后赋值给上一个item
        if previous is not None:
            fourWidgetData = self.getPrettyDataInSameYTTW_ExcelDataTW_ExcelFIleAddressLE_YTCB()
            previous.setData(Qt.UserRole, fourWidgetData)
        # 再把现在这个item的数据放置在五个组件上
        if current is not None:
            PrettyData = current.data(Qt.UserRole)
            self.setPrettyDataToSameYTTW_ExcelDataTW_ExcelFIleAddressLE_exceptYTCBs_onlyYTCBs(PrettyData)
        else:
            self.clearNineWidgetData()

    @pyqtSlot()
    def clearNineWidgetData(self):
        self.ui.excelFileAddressLE.clear()
        self.ui.excelDataTW.clearContents()
        self.ui.sameYTTW.clearContents()
        self.ui.onlyYT1CB.clear()
        self.ui.onlyYT2CB.clear()
        self.ui.onlyYT3CB.clear()
        self.ui.exceptYT1CB.clear()
        self.ui.exceptYT2CB.clear()
        self.ui.exceptYT3CB.clear()

    def getPrettyDataInSameYTTW_ExcelDataTW_ExcelFIleAddressLE_YTCB(self):
        # 从四个组件中获取数据 然后设置在List的Data中
        fileAddress = self.getExcelFileAddressLEText()
        sortedData = self.getDataInExcelDataTW()
        sameYTCount = self.getDataInSameYTTW()
        ytSet = self.getCBYTData()
        sortedData_sameYTCount_fileAddress_ytSet = {"sortedData": sortedData, "sameYTCount": sameYTCount,
                                                    "ytSet": ytSet, "fileAddress": fileAddress}
        return sortedData_sameYTCount_fileAddress_ytSet

    def getExcelFileAddressLEText(self):
        # 获取excelFileAddressLE的文本
        fileAddress = self.ui.excelFileAddressLE.text()
        print("在excelFileAddressLE中获取到以下数据列表" + str(fileAddress))
        return fileAddress

    def getDataInExcelDataTW(self):
        # 遍历整个tableWidget 然后存成List
        tableWidgetRow = self.ui.excelDataTW.rowCount()
        tableWidgetColumn = self.ui.excelDataTW.columnCount()
        excelDataTWValueList = [[] for i in range(tableWidgetRow)]
        # 创建一个空的二维列表 然后用坐标的方式进行填充
        for row in range(tableWidgetRow):
            for column in range(tableWidgetColumn):
                itemData = self.ui.excelDataTW.item(row, column)
                excelDataTWValueList[row].append(itemData.text())
        print("从ExcelDataTW中获取到以下数据列表：" + str(excelDataTWValueList))
        return excelDataTWValueList

    def getDataInSameYTTW(self):
        # 同样是获取二维数据列表
        tableWidgetRow = self.ui.sameYTTW.rowCount()
        tableWidgetColumn = self.ui.sameYTTW.columnCount()
        sameYTTWValueList = [[] for i in range(tableWidgetRow)]
        # 创建一个空的二维列表 然后用坐标的方式进行填充
        for row in range(tableWidgetRow):
            for column in range(tableWidgetColumn):
                itemData = self.ui.sameYTTW.item(row, column)
                sameYTTWValueList[row].append(itemData.text())
        print("从sameYTTW中获取到以下数据列表：" + str(sameYTTWValueList))
        return sameYTTWValueList

    def getCBYTData(self):
        # 因为六个YTCB都是同步的 所以只需要取一个CB的数值列表就可以了
        YTCBTextList = []
        comboBoxCount = self.ui.onlyYT1CB.count()
        for index in range(comboBoxCount):
            comboBoxIndexText = self.ui.onlyYT1CB.itemText(index)
            YTCBTextList.append(comboBoxIndexText)
        print("在onlyYT1CB中获取到以下数据列表" + str(YTCBTextList))
        return YTCBTextList

    def setPrettyDataToSameYTTW_ExcelDataTW_ExcelFIleAddressLE_exceptYTCBs_onlyYTCBs(self, dataDict):
        # 接受字典 字典中有三=四个key分别是fileAddress,sortedData,sameYTCount,ytSet
        # 设置两个大组件 七个小组件
        # 从字典中取出四个数据
        fileAddress = dataDict['fileAddress']
        sortedData = dataDict['sortedData']
        sameYTCount = dataDict['sameYTCount']
        ytSet = dataDict['ytSet']
        if len(ytSet) and ytSet[0] != "":  # 如果月台集合第一位不是空值且yeSet不为空则
            ytSet.insert(0, "")  # 对第一位插入为空值，以便用户可选三个月台以下的操作
        self.setExcelFileAddressLEText(fileAddress)
        self.setDataToExcelDataTW(sortedData)
        self.setDataToSameYTTW(sameYTCount)
        self.setOnlyCBYTData(ytSet)
        self.setExceptCBYTData(ytSet)

    def setOnlyCBYTData(self, ytSetData):
        # 设置三个onlyCB的月台数据
        print("清空三个onlyYTCB组合框组件后再填充以下数据" + str(ytSetData))
        self.ui.onlyYT1CB.clear()
        self.ui.onlyYT2CB.clear()
        self.ui.onlyYT3CB.clear()
        self.ui.onlyYT1CB.addItems(ytSetData)
        self.ui.onlyYT2CB.addItems(ytSetData)
        self.ui.onlyYT3CB.addItems(ytSetData)
        self.ui.onlyYT1CB.setCurrentIndex(1)

    def setExceptCBYTData(self, ytSetData):
        # 设置三个exceptCB的月台数据
        print("清空三个exceptYTCB组合框组件后再填充以下数据" + str(ytSetData))
        self.ui.exceptYT1CB.clear()
        self.ui.exceptYT2CB.clear()
        self.ui.exceptYT3CB.clear()
        self.ui.exceptYT1CB.addItems(ytSetData)
        self.ui.exceptYT2CB.addItems(ytSetData)
        self.ui.exceptYT3CB.addItems(ytSetData)
        self.ui.exceptYT1CB.setCurrentIndex(1)

    def setExcelFileAddressLEText(self, fileAddress):
        # 本来是不打算写这个函数的，但后来想一下log中没有设置地址的信息就很难受
        print("将要向ExcelFileAddressLE中填充以下数据：" + str(fileAddress))
        self.ui.excelFileAddressLE.setText(fileAddress)

    def setDataToExcelDataTW(self, ListData):
        # 设置数据到excelDataTW表格组件中
        self.ui.excelDataTW.clearContents()  # 刷新
        print("清空ExcelDataTW组件后再填充以下数据：" + str(ListData))
        self.ui.excelDataTW.setRowCount(len(ListData))
        for oneDimensionalIndex, oneDimensionalDataList in enumerate(ListData):  # 用enumrate的方法同时获取坐标和数据
            for twoDimensionalIndex, twoDimensionalDataList in enumerate(oneDimensionalDataList):
                aItem = QTableWidgetItem(str(twoDimensionalDataList))
                aItem.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)  # 设置项目垂直水平对齐
                self.ui.excelDataTW.setItem(oneDimensionalIndex, twoDimensionalIndex, aItem)
        self.ui.excelDataTW.resizeColumnsToContents()  # 调整列宽

    def setDataToSameYTTW(self, ListData):
        # 设置数据到sameYTTW表格组件中 两个都是设置二位列表 那么写法都一样了
        self.ui.sameYTTW.clearContents()  # 刷新
        print("清空SameYTTW组件后再填充以下数据：" + str(ListData))
        self.ui.sameYTTW.setRowCount(len(ListData))
        for oneDimensionalIndex, oneDimensionalDataList in enumerate(ListData):  # 用enumrate的方法同时获取坐标和数据
            for twoDimensionalIndex, twoDimensionalDataList in enumerate(oneDimensionalDataList):
                aItem = QTableWidgetItem(str(twoDimensionalDataList))
                aItem.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)  # 设置项目垂直水平对齐
                self.ui.sameYTTW.setItem(oneDimensionalIndex, twoDimensionalIndex, aItem)

    @pyqtSlot()
    def on_getfilePB_clicked(self):
        # 获取文件名列表
        curPath = os.path.join(os.path.expanduser("~"), "Desktop")
        dlgTitle = "请选择要处理的Excel"
        filt = "文档(*.xlsx *.xls);;所有(*.*)"
        fileAddresslist, filtUsed = QFileDialog.getOpenFileNames(self, dlgTitle, curPath, filt)
        # 如果列表不为空，则添加漂亮的数据到暂存表中
        if fileAddresslist:
            print("待处理的Excel文件地址列表为:" + str(fileAddresslist))
            self.AddPrettyDataToExcelFileListWidget(fileAddresslist)

    def AddPrettyDataToExcelFileListWidget(self, fileAddressList=None):
        # 在这里调用函数处理excel文件 并把函数返回值添加到listwidget中
        aItem = None
        for fileAddress in fileAddressList:  # 逐个处理每个文件
            fileName = os.path.basename(fileAddress)  # 取基础文件名
            inExcelFileLWaddedFileNameList = self.getInExcelFileLWaddedFileNameList()
            print("获取到已经添加到ExcelFileLW组件中的文件名列表为：" + str(inExcelFileLWaddedFileNameList))

            if fileName not in inExcelFileLWaddedFileNameList:  # 如果这个文件名不在已添加的文件名列表中
                prettyListData = self.handleExcelFileToPretty(fileAddress)
                if not prettyListData:
                    continue
                aItem = QListWidgetItem()  # 设置初始化项目
                aItem.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)  # 设置项目垂直水平对齐
                aItem.setText(fileName)  # 设置项目名
                aItem.setData(Qt.UserRole, prettyListData)  # 设置项目数据
                self.ui.excelFileLW.addItem(aItem)  # 向组件中添加项目
            else:
                inaItemDatafileAddress = self.getInExcelFileLWaddedAppointFileAdress(fileName)  # 否则打印出找到这个同名文件地址
                print("已存在于列表中的同名文件绝对地址为：" + str(inaItemDatafileAddress))
                print("当前欲添加重复文件名称的绝对文件地址为：" + str(fileAddress))
                longlengthtext = "文件名重复,如要替换请先删除列表中的同名数据\n已存在于列表中的同名文件绝对地址为：\n" \
                                 + str(inaItemDatafileAddress) + "\n当前重复文件名称的绝对文件地址为：\n" + str(fileAddress)
                QMessageBox.information(self, "提示", longlengthtext)
                continue

        # 最后一个数据添加完后 就要激活其中一个数据，默认是最后一个也就是最新追加的一个   注意这次更新没有通过项目双击信号
        setCurrentRow(row)
        # if aItem is not None:
        #     self.ui.excelFileLW.setCurrentItem(aItem)
        #     print("设置ExcelFileLW的当前项目为:" + str(aItem))



    def getInExcelFileLWaddedFileNameList(self):
        # 获取已经添加到excelFileLW组件中的基础文件名列表 并返回
        inExcelFileLWaddedFileNameList = []
        for index in range(self.ui.excelFileLW.count()):
            inExcelFileLWaddedFileNameList.append(self.ui.excelFileLW.item(index).text())
        return inExcelFileLWaddedFileNameList

    def handleExcelFileToPretty(self, fileAddress=None):
        # 这个函数只负责可以成功打开文件并获取到数据 具体数据处理再下一层
        # 如果文件地址不为空
        if fileAddress:
            print("本次处理的Excel文件地址为:" + str(fileAddress))
            # 获取四个指定行列
            shipidCID = self.ui.shipCIDLE.text()  # 单号数据列 column
            scantimeCID = self.ui.scanCIDLE.text()  # 扫描时间列
            ytCID = self.ui.ytCIDLE.text()  # 月台号列
            effectivedataRID = self.ui.effectiveRIDLE.text()  # 第一行有效数据 row

            print("获取到单号列为:" + str(shipidCID))
            print("获取到扫描时间列为:" + str(scantimeCID))
            print("获取到月台号列为:" + str(ytCID))
            print("获取到第一行有效数据行为:" + str(effectivedataRID))

            try:
                # 加载excel文件
                wb = load_workbook(fileAddress)
                ws = wb.active  # 激活这个workbook
                shipIdData = ws[shipidCID]  # 单号列数据
                scanTimeData = ws[scantimeCID]  # 扫描时间列数据
                ytData = ws[ytCID]  # 月台号列数据

                filename = os.path.basename(fileAddress)  # 获取文件名称
                print("文件名为：" + filename)
                print(shipidCID + "列表头为：" + str(shipIdData[0].value))
                print(scantimeCID + "列表头为：" + str(scanTimeData[0].value))
                print(ytCID + "列表头为：" + str(ytData[0].value))
                print("注意审查表头是否正确")
                # 传入的列号 是为了更好的打印行列中有None的信息
                rawData = self.filteredDataContainsNone(shipidCID, scantimeCID, ytCID, effectivedataRID,
                                                        shipIdData, scanTimeData, ytData)
                sortedData, sameYTCount, ytSet = self.handleExcelTwoDimensionalData(rawData)  # 处理excel二维列表数据
                sortedData_sameYTCount_fileAddress_ytSet = {"sortedData": sortedData, "sameYTCount": sameYTCount,
                                                            "ytSet": ytSet, "fileAddress": fileAddress}
                return sortedData_sameYTCount_fileAddress_ytSet
            except InvalidFileException:
                QMessageBox.warning(self, "警告", "文件格式错误")
                print("文件格式错误" + str(fileAddress))
                return
            except FileNotFoundError:
                QMessageBox.warning(self, "警告", "文件查找错误")
                print("文件查找错误" + str(fileAddress))
                return
            # except Exception as e:
            #     # print("未知错误", e)
            #     logger.error("未知错误")
            #     logger.error(e)



    def handleExcelTwoDimensionalData(self, tableData):
        ytData = [ytStr[2] for ytStr in tableData]
        ytSet = set()  # 月台集合
        [ytSet.add(str(i)) for i in ytData]  # 收集月台集合

        # 从这里开始进行排序处理 主要是不同月台间的时间排序
        # 我需要先把相同月台的单独拆出来 排好时间再组合回来
        tableData.sort(key=lambda elem: elem[2])
        sortedYTData = tableData[:]
        print("仅有月台排序的数据：" + str(sortedYTData))

        # element2 也就是月台 先排序月台
        ytSet = sorted(list(ytSet))  # 转换月台集合为列表并排序
        # ytSet.sort()
        print("原始月台集合：" + str(ytSet))
        # 从这里开始排序时间
        sameYTData = [[] for i in range(len(ytSet))]
        # 用月台集合创建一个空的二位列表 方便填充
        [
            sameYTData[j].append(sortedYTData[i])
            for j in range(len(ytSet))
            for i in range(len(sortedYTData))
            if sortedYTData[i][2] == ytSet[j]
            # 相同的月台都加到独立列表中 这步主要是为了时间排序
        ]
        # 时间排序 注意sameYTData的结构 他是三维的 我这里用for解开了一层之后再用的匿名函数
        [i.sort(key=lambda elem: elem[1]) for i in sameYTData]
        sortedYTAndTimeData = [j for i in sameYTData for j in i]
        # 即月台和时间都排好顺序了 把这个三维重组成二维

        # 从这里开始对相同月台数量进行统计
        sameYTCount = [[i, 0] for i in ytSet]
        # 列表中每个元素都初始为0个
        # [[[312102088273409, '2021-12-07 01:58:00', 'YT212']],
        # [[4280383234896, '2021-12-06 20:34:00', 'YT502']],
        # [[[312102088273409, '2021-12-07 01:58:00', 'YT212']
        for x in sameYTData:
            for y in x:  # x也是一个列表 上面就是sameYTData的示例
                for z in sameYTCount:
                    if y[2] == z[0]:  # 这个统计方式是很经典的一种
                        z[1] += 1
        sameYTCount.sort(key=lambda elem: elem[1], reverse=True)
        # 对数量进行排序倒序 最大的排上面
        print("月台和时间都排列好顺序的数据：" + str(sortedYTAndTimeData))
        print("数量倒序的相同月台列表：" + str(sameYTCount))
        return sortedYTAndTimeData, sameYTCount, ytSet


if __name__ == "__main__":  # 用于当前窗体测试
    from app.esheet_process_widget.UI.ui_ExcelProcess import Ui_EPWC

    app = QApplication(sys.argv)  # 创建GUI应用程序
    form = QmyExcelProcess()  # 创建窗体

    form.show()
    sys.exit(app.exec_())
else:
    from app.esheet_process_widget.UI.ui_ExcelProcess import Ui_EPWC


# 关于event的accept
# 在书中5.1中有说明 例如 closeevent的event参数实际上是QCloseEvent类型
# 也就是注意这个默认处理函数和QEvent的类型 TYPE
# 对于关闭实际关闭窗口的是QCloseEvent的accept函数
