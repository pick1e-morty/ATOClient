from pathlib import Path
import sys
from PyQt5.QtWidgets import QWidget, QApplication
from loguru import logger
from typing import Union, Dict

from qfluentwidgets import MessageBox, Dialog

from app.esheet_process_widget.epw_define import YTCTFEnum, YTConfigDataStruct
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.utils.tools import is_instance_variables_has_empty, is_empty


def getValidDevType(devType: str) -> str:
    # 从表里拿到的设备类型转成英文的，转不成就返回None
    validDevType = None
    if devType == "大华":
        validDevType = "dahua"
    elif devType == "海康":
        validDevType = "haikang"
    return validDevType


def isValidIPv4(ip: str) -> bool:
    # 判断字符串是否为合法ipv4
    # 将字符串按照点号分割成四个部分
    parts = ip.split(".")
    if len(parts) != 4:
        return False
    # 判断每个数字是否在0到255之间
    for part in parts:
        try:
            num = int(part)
            if num < 0 or num > 255:
                return False
        except ValueError:
            return False
    # 所有数字都符合条件，说明该字符串是合法的IPv4地址
    return True


class DevConfigFileNotFoundError(Exception):
    """
    自定义异常类，用于标记设备配置文件不存在的情况
    """

    def __init__(self, message="设备配置文件不存在，整个文件路径最好是英文的且绝对不能带有空格"):
        self.message = message
        super().__init__(self.message)


class DevConfigFileInvalidError(Exception):
    """
    自定义异常类，用于标记设备配置文件格式不正确的情况
    """

    def __init__(self, message="设备配置文件格式错误，整个文件路径最好是英文的且绝对不能带有空格"):
        self.message = message
        super().__init__(self.message)


class DevConfigFileContentIsEmptyException(Exception):
    """
    自定义异常类，用于标记配置文件是空的情况
    """

    def __init__(self, message="设备配置文件为空"):
        self.message = message
        super().__init__(self.message)


class DevConfigFileHandleErrorException(Exception):
    """
    自定义异常类，用于标记处理配置文件出错的情况
    """

    def __init__(self, message="设备配置文件处理错误"):
        self.message = message
        super().__init__(self.message)


class DevConfigFileContentIsInvalidException(Exception):
    """
    自定义异常类，用于标记配置文件中的数值有误的情况，
    """

    def __init__(self, message="设备配置文件数值错误"):
        self.message = message
        super().__init__(self.message)


def getYtConfigFromConfigFile(devConfigFilePath: Union[Path, str] = None) -> Dict[str, YTConfigDataStruct]:
    # 从配置文件中读取月台配置
    devConfigFilePath = Path(__file__).parent.parent / "AppData/月台配置.xlsx" if devConfigFilePath is None else devConfigFilePath
    # 如果用户没传地址，就用缺省的
    try:
        wb = load_workbook(str(devConfigFilePath))  # 打开配置文件
    except InvalidFileException:
        raise DevConfigFileInvalidError(f"设备配置文件格式错误或文件路径有误，确保文件格式为xlsx或xls及文件路径中不能含有空格\n{str(devConfigFilePath)}")
    except FileNotFoundError:
        raise DevConfigFileNotFoundError(f"设备配置文件不存在，确保文件格式为xlsx或xls及文件路径中不能含有空格\n{str(devConfigFilePath)}")
    ws = wb.active  # 获取活动表，如果用户改变了默认的激活表，那这里也是会出错的地方，到时候会报配置文件为空
    ws.delete_rows(1)  # 删除第一行标题
    ytList = [cell.value for cell in ws[YTCTFEnum.YT.value]]  # 月台列数据
    devTypeList = [cell.value for cell in ws[YTCTFEnum.devType.value]]  # 设备类型列数据
    devIPList = [cell.value for cell in ws[YTCTFEnum.devIP.value]]  # 设备IP列数据
    devChannelList = [cell.value for cell in ws[YTCTFEnum.devChannel.value]]  # 设备通道列数据
    devPortList = [cell.value for cell in ws[YTCTFEnum.devPort.value]]  # 设备端口列数据
    devUserNameList = [cell.value for cell in ws[YTCTFEnum.devUserName.value]]  # 设备用户名列数据
    devPasswordList = [cell.value for cell in ws[YTCTFEnum.devPassword.value]]  # 设备密码列数据
    wb.close()  # 用完就关
    ytConfigList = {}

    devConfigFileContentInvalidText = ""  # 个别行的配置有问题，统计到一块然后raise给上层

    for index in range(len(ytList)):
        temp_ytConfig = YTConfigDataStruct()
        try:
            temp_ytConfig.yt = str(ytList[index])
            if is_empty(temp_ytConfig.yt):  # 如果月台名称这个单元格是空的，一般说明用户删除了表格数据，但xlsx仍会保存这个单元格的数据为None，提前跳出下
                continue

            devTypeResult = getValidDevType(str(devTypeList[index]))  # 成功返回正确的英文设备类型，不然返回的None，这个判断用了两行
            temp_ytConfig.devType = devTypeResult if devTypeResult is not None else None

            temp_ytConfig.devIP = str(devIPList[index]) if isValidIPv4(str(devIPList[index])) else None
            # int方法不能接None，所以要提前判断一下，如果是None就让is_instance_variables_has_empty去处理
            temp_ytConfig.devChannel = int(devChannelList[index]) if str(devChannelList[index]).isnumeric() else None
            temp_ytConfig.devPort = int(devPortList[index]) if str(devPortList[index]).isnumeric() else None
            temp_ytConfig.devUserName = str(devUserNameList[index])
            temp_ytConfig.devPassword = str(devPasswordList[index])
            if is_instance_variables_has_empty(temp_ytConfig):  # 上面的数值判断，如果有错误的数值就会被设置为None，记录原数值后统一raise给上层
                errorText = f"月台号[{ytList[index]}],设备厂商[{devTypeList[index]}],IP[{devIPList[index]}],通道序号[{devChannelList[index]}],端口[{devPortList[index]}],用户名[{devUserNameList[index]}],密码[{devPasswordList[index]}]\n"
                devConfigFileContentInvalidText += errorText
            else:
                ytConfigList[temp_ytConfig.yt] = temp_ytConfig
        except IndexError as errorInfo:  # 某个单元格为空，但应该不会存在这个问题的，先抓了再说吧
            logger.error(f"好像是未知错误，原报错信息{errorInfo}")
            raise DevConfigFileHandleErrorException(f"未知错误，原报错信息\n{errorInfo}")
        except ValueError as errorInfo:  # 配置表中数值有误，应该是填错位置了
            logger.error(f"好像是未知错误，原报错信息{errorInfo}")
            raise DevConfigFileHandleErrorException(f"未知错误，原报错信息\n{errorInfo}")

    if devConfigFileContentInvalidText != "":
        title = "检测到配置文件中有以下行中参数有误\n"
        devConfigFileContentInvalidText = title + devConfigFileContentInvalidText
        raise DevConfigFileContentIsInvalidException(devConfigFileContentInvalidText)

    if ytConfigList:  # 判断列表是否为空
        return ytConfigList
    else:
        raise DevConfigFileContentIsEmptyException(f"设备配置文件是空的，注意检查是否填错格式\n{devConfigFilePath}")


def YtBindDevConfigGenerate():
    # 这是一个生成器，用于获取月台绑定的设备配置信息
    # 上层用send发送月台索引到这，查到之后把配置yield出去
    ytBindConfigDict = getYtConfigFromConfigFile()  # 读取设备配置表
    while True:
        ytName = yield
        ytConfig = ytBindConfigDict.get(ytName, None)
        yield ytConfig
